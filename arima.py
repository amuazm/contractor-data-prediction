# ARIMA

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from statsmodels.tsa.arima.model import ARIMA
from dateutil import relativedelta
from openpyxl import load_workbook

# Load Result.xlsx
wb_result = load_workbook("./Files/Output/Result.xlsx")
ws_budget = wb_result["Budget"]
ws_reports = wb_result["Reports"]

# Add "Remark" column
ws_reports.insert_cols(3, 1)
ws_reports["C1"] = "Remark"
for row in ws_reports.iter_rows(min_row=2):
    row[2].value = "Actual Date"

# Get Overall Budgets
overall_budgets = {}
for row in ws_budget.iter_rows(min_row=2):
    overall_budgets[row[0].value] = row[4].value
# Get Durations
durations = {}
for row in ws_budget.iter_rows(min_row=2):
    durations[row[0].value] = row[2].value
# Get Monthly Budgets
monthly_budgets = {}
for row in ws_budget.iter_rows(min_row=2):
    monthly_budgets[row[0].value] = row[3].value

# Get Reports sheet as dataframe for input into ARIMA function
df_reports = pd.read_excel("./Files/Output/Result.xlsx", sheet_name="Reports")

# Perform ARIMA for each project
for project_id in overall_budgets:
    print("\n\n\n\n\n\n==========================", project_id, "==========================")
    # if project_id != "Project 235":
    #     continue
    # Filter project ID
    df = df_reports.loc[df_reports["Project ID"] == project_id]
    # Filter Date (Time-Series), ACWP, and BCWP (forecasting these 2 metrics)
    df = df[["Date", "ACWP", "BCWP"]]
    df = df.set_index(["Date"])
    df = df.diff().fillna(df)
    df.index = pd.to_datetime(df.index) - pd.tseries.offsets.MonthBegin(1)
    df = df.asfreq(pd.infer_freq(df.index))
    # plt.plot(df.index, df["ACWP"])
    # plt.plot(df.index, df["BCWP"])
    # plt.show()

    months_passed = len(df)
    project_duration = durations[project_id]
    planned_months_left = project_duration - months_passed
    estimated_month_variance = df_reports.loc[df_reports["Project ID"] == project_id]["VAC(t)"].iloc[-1] # Negative means extra months estimated (behind schedule)
    months_to_predict = np.ceil(durations[project_id] - len(df) - df_reports.loc[df_reports["Project ID"] == project_id]["VAC(t)"].iloc[-1])

    pred_start_date = df.index[-1:]
    pred_start_date = pred_start_date.to_pydatetime()[0] + relativedelta.relativedelta(months=1)
    pred_end_date = pred_start_date + relativedelta.relativedelta(months=months_to_predict)

    lags = len(df.index)//3

    # Split ACWP and BCWP
    l = []
    l.append(df.drop(["BCWP"], axis=1))
    l.append(df.drop(["ACWP"], axis=1))

    l2 = []
    l3 = []
    try:
        for df2 in l:
            # pacf_plot = plot_pacf(df2, lags=lags, method="ywm")
            # plt.show()

            #TODO: I MA
            model = ARIMA(df2, order=(lags, 0, 0))
            model_fit = model.fit()

            predictions = model_fit.predict(start=pred_start_date, end=pred_end_date)
            predictions = predictions.to_frame()
            predictions.index.name = "Date"
            predictions.columns = [list(df2)[0]]

            df2 = df2.cumsum()
            l3.append(df2)

            predictions_insert = df2.iloc[-1].to_frame().T
            predictions_insert.index.name = "Date"
            predictions = pd.concat([predictions_insert, predictions])
            predictions = predictions.cumsum()
            predictions = predictions.iloc[1:, :]

            l2.append(predictions)

        df = pd.concat([l2[0], l2[1]], axis=1)

        reached_one_hundred_percent = False
        for index, row in df.iterrows():
            if reached_one_hundred_percent == False:
                if row["BCWP"] >= overall_budgets[project_id]:
                    row["BCWP"] = overall_budgets[project_id]
                    reached_one_hundred_percent = True
                ws_reports.append([project_id, index.date(), "Forecasted", "", "", row["ACWP"], row["BCWP"]])
    except Exception as e:
        print(e)

d = {}
for row in ws_reports.iter_rows(min_row=2):
    # Completion Percentage
    if row[3].value == "":
        row[3].value = row[6].value / overall_budgets[row[0].value]

    # Months Passed
    if row[0].value not in d:
        d[row[0].value] = 1
    else:
        d[row[0].value] += 1
    if row[4].value == "":
        row[4].value = d[row[0].value]

    # BCWS
    if row[7].value == None:
        bcws = monthly_budgets[row[0].value] * row[4].value
        if bcws >= overall_budgets[row[0].value]:
            bcws = overall_budgets[row[0].value]
        row[7].value = bcws
    # CPI
    if row[8].value == None:
        row[8].value = row[6].value / row[5].value
    # CV
    if row[9].value == None:
        row[9].value = row[6].value - row[5].value
    # SPI
    if row[10].value == None:
        row[10].value = row[6].value / row[7].value
    # SV
    if row[11].value == None:
        row[11].value = row[6].value - row[7].value
    # EAC
    if row[12].value == None:
        row[12].value = row[5].value + (overall_budgets[row[0].value] - row[6].value) / row[8].value
    # EAC(t)
    if row[13].value == None:
        row[13].value = row[4].value + (max(durations[row[0].value], row[4].value) - row[4].value * row[10].value) / row[10].value
    # VAC
    if row[14].value == None:
        row[14].value = overall_budgets[row[0].value] - row[12].value
    # VAC(t)
    if row[15].value == None:
        row[15].value = durations[row[0].value] - row[13].value

    row[3].style = "Percent"
    row[5].style = "Currency"
    row[6].style = "Currency"
    row[7].style = "Currency"
    row[9].style = "Currency"
    row[11].style = "Currency"
    row[12].style = "Currency"
    row[14].style = "Currency"

# Save as Result_ARIMA.xlsx
wb_result.save("./Files/Output/Result_ARIMA.xlsx")