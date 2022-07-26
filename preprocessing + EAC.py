from openpyxl import load_workbook
import shutil

# Copy file
src = "./Files/Input/Contractor Project Management Dataset.xlsx"
dest = "./Files/Output/Result.xlsx"
shutil.copyfile(src, dest)

# Result WB and WS
wb_result = load_workbook("./Files/Output/Result.xlsx")
ws_reports = wb_result["Reports"]
ws_budget = wb_result["Budget"]

# Rename ACWP
ws_reports["D1"].value = "ACWP"

# Months Passed
ws_reports.insert_cols(4)
ws_reports["D1"] = "Months Passed"
i = 1
current_project = ""
for row in ws_reports.iter_rows(min_row=2):
    if row[0].value != current_project:
        current_project = row[0].value
        i = 1
    row[3].value = i
    i += 1

# BCWP Column
ws_reports.insert_cols(6)
ws_reports["F1"].value = "BCWP"
# Get Overall Budgets
overall_budgets = {}
for row in ws_budget.iter_rows(min_row=2):
    overall_budgets[row[0].value] = row[4].value
# monthly_budgets for later down the line
monthly_budgets = {}
for row in ws_budget.iter_rows(min_row=2):
    monthly_budgets[row[0].value] = row[3].value
# Multiply
for row in ws_reports.iter_rows(min_row=2):
    row[5].value = overall_budgets[row[0].value] * row[2].value
    row[5].style = "Currency"

# CPI
ws_reports.insert_cols(8)
ws_reports["H1"].value = "CPI"
# Multiply
for row in ws_reports.iter_rows(min_row=2):
    row[7].value = row[5].value / row[4].value

# CV
ws_reports.insert_cols(9)
ws_reports["I1"].value = "CV"
# Minus
for row in ws_reports.iter_rows(min_row=2):
    row[8].value = row[5].value - row[4].value
    row[8].style = "Currency"

# SPI
ws_reports.insert_cols(10)
ws_reports["J1"].value = "SPI"
# Multiply
for row in ws_reports.iter_rows(min_row=2):
    row[9].value = row[5].value / row[6].value

# SV
ws_reports.insert_cols(11)
ws_reports["K1"].value = "SV"
# Multiply
for row in ws_reports.iter_rows(min_row=2):
    row[10].value = row[5].value - row[6].value
    row[10].style = "Currency"

# EAC
ws_reports.insert_cols(12)
ws_reports["L1"].value = "EAC"
# EAC = ACWP + (Overall Budget - BCWP) / 1
for row in ws_reports.iter_rows(min_row=2):
    row[11].value = row[4].value + (overall_budgets[row[0].value] - row[5].value)
    row[11].style = "Currency"

# EAC/CPI
ws_reports.insert_cols(13)
ws_reports["M1"].value = "EAC/CPI"
for row in ws_reports.iter_rows(min_row=2):
    row[12].value = row[4].value + (overall_budgets[row[0].value] - row[5].value) / row[7].value
    row[12].style = "Currency"

# EAC/SPI
ws_reports.insert_cols(14)
ws_reports["N1"].value = "EAC/SPI"
for row in ws_reports.iter_rows(min_row=2):
    row[13].value = row[4].value + (overall_budgets[row[0].value] - row[5].value) / row[9].value
    row[13].style = "Currency"

# EAC/SCI
ws_reports.insert_cols(15)
ws_reports["O1"].value = "EAC/SCI"
for row in ws_reports.iter_rows(min_row=2):
    row[14].value = row[4].value + (overall_budgets[row[0].value] - row[5].value) / (row[7].value * row[9].value)
    row[14].style = "Currency"

# EAC(t)
ws_reports.insert_cols(16)
ws_reports["P1"].value = "EAC(t)"
# Get Durations
durations = {}
for row in ws_budget.iter_rows(min_row=2):
    durations[row[0].value] = row[2].value
# EAC(t) = Duration (Months) - TV | TV = SV / PVrate | PVrate = Overall Budget / Duration (Months)
for row in ws_reports.iter_rows(min_row=2):
    row[15].value = durations[row[0].value] - (row[10].value / (overall_budgets[row[0].value] / durations[row[0].value]))

# EAC(t)/SPI
ws_reports.insert_cols(17)
ws_reports["Q1"].value = "EAC(t)/SPI"
for row in ws_reports.iter_rows(min_row=2):
    row[16].value = row[15].value / row[9].value

# EAC(t)/SCI
ws_reports.insert_cols(18)
ws_reports["R1"].value = "EAC(t)/SCI"
for row in ws_reports.iter_rows(min_row=2):
    row[17].value = row[15].value / (row[7].value * row[9].value)

# EAC(t)(ED)
ws_reports.insert_cols(19)
ws_reports["S1"].value = "EAC(t)(ED)"
# EAC(t)(ED) = Months Passed + (max(Duration (Months), Months Passed) - ED) / PF | ED = Months Passed * SPI
for row in ws_reports.iter_rows(min_row=2):
    row[18].value = row[3].value + (max(durations[row[0].value], row[3].value) - row[3].value * row[9].value)

# EAC(t)(ED)/SPI
ws_reports.insert_cols(20)
ws_reports["T1"].value = "EAC(t)(ED)/SPI"
for row in ws_reports.iter_rows(min_row=2):
    row[19].value = row[3].value + (max(durations[row[0].value], row[3].value) - row[3].value * row[9].value) / row[9].value

# EAC(t)(ED)/SCI
ws_reports.insert_cols(21)
ws_reports["U1"].value = "EAC(t)(ED)/SCI"
for row in ws_reports.iter_rows(min_row=2):
    row[20].value = row[3].value + (max(durations[row[0].value], row[3].value) - row[3].value * row[9].value) / (row[7].value * row[9].value)

# Remove Unnecessary Columns for Simplicity
ws_reports.delete_cols(12, 1)
ws_reports.delete_cols(13, 2)
ws_reports.delete_cols(13, 1)
ws_reports.delete_cols(13, 3)
ws_reports.delete_cols(14, 1)

# # ETC
# ws_reports.insert_cols(14)
# ws_reports["N1"].value = "ETC"
# # ETC = EAC - ACWP
# for row in ws_reports.iter_rows(min_row=2):
#     row[13].value = row[11].value - row[4].value
#     row[13].style = "Currency"

# VAC
ws_reports.insert_cols(14)
ws_reports["N1"].value = "VAC"
# VAC = Overall Budget - EAC
for row in ws_reports.iter_rows(min_row=2):
    row[13].value = overall_budgets[row[0].value] - row[11].value
    row[13].style = "Currency"

# VAC(t)
ws_reports.insert_cols(15)
ws_reports["O1"].value = "VAC(t)"
# VAC(t) = Duration (Months) - EAC(t)
for row in ws_reports.iter_rows(min_row=2):
    row[14].value = durations[row[0].value] - row[12].value

# # TCPI
# ws_reports.insert_cols(16)
# ws_reports["P1"].value = "TCPI"
# # TCPI = (Overall Budget - BCWP)/(Overall Budget - ACWP)
# for row in ws_reports.iter_rows(min_row=2):
#     row[15].value = (overall_budgets[row[0].value] - row[5].value)/(overall_budgets[row[0].value] - row[4].value)

wb_result.save("./Files/Output/Result.xlsx")